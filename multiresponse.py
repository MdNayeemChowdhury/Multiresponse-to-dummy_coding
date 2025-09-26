import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font
import streamlit as st

st.title("📊 Multi-Response Dummy Coding")

# Step 1: Upload file
uploaded_file = st.file_uploader("Upload CSV or Excel file", type=["csv", "xlsx"])

if uploaded_file:
    # Step 2: Load file
    if uploaded_file.name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)

    # Step 3: Select multi-response variables
    multi_vars = st.multiselect("Select the multi-response variable(s)", df.columns.tolist())

    # Step 4: Separator input
    sep = st.text_input(
        "Enter the separator used in multi-response variables",
        value=","
    )

    # Step 5: Option to keep or drop originals
    keep_original = st.checkbox("Keep original multi-response columns in final dataset?", value=True)

    # Step 6: Run transformation
    if st.button("Run Dummy Coding"):
        for var in multi_vars:
            df[var] = df[var].fillna("").astype(str)
            unique_responses = sorted(set(
                x.strip() for row in df[var] for x in row.split(sep) if x.strip()
            ))

            st.write(f"**Unique responses in {var}:** {unique_responses}")

            for resp in unique_responses:
                df[f"{var}_{resp}"] = df[var].apply(
                    lambda x: 1 if resp in [r.strip() for r in x.split(sep)] else 0
                )

        # Drop originals if chosen
        if not keep_original:
            df = df.drop(columns=multi_vars)

        # Step 7: Save to Excel in memory
        output = BytesIO()
        df.to_excel(output, index=False)
        output.seek(0)

        # Step 8: Apply formatting
        wb = load_workbook(output)
        ws = wb.active

        # Bold headers
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Autofit columns
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Save formatted file
        formatted_output = BytesIO()
        wb.save(formatted_output)
        formatted_output.seek(0)

        # Step 9: Download button
        st.success("✅ Dummy coding completed!")
        st.download_button(
            label="📥 Download Dummy-Coded Dataset",
            data=formatted_output,
            file_name="dummy_coded_dataset.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# --- Footer Section ---
st.markdown(
    """
    <style>
        .footer {
            position: fixed;
            left: 0;
            bottom: 0;
            width: 100%;
            background-color: #007BFF;  /* Full blue bar */
            color: white;
            text-align: center;
            padding: 12px 0;
            font-size: 14px;
        }
        .footer a {
            color: white;
            text-decoration: underline;
        }
    </style>

    <div class="footer">
        Developed by <b>Md. Nayeem Chowdhury</b> | 
        Contact: <a href="mailto:mnchowdhury@isrt.ac.bd">mnchowdhury@isrt.ac.bd</a>
    </div>
    """,
    unsafe_allow_html=True
)
